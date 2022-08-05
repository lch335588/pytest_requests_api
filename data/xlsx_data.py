import ast
import openpyxl
import requests

from config.config import *
import re
import jsonpath


class XlsxData:
    def __init__(self):
        self.wb = openpyxl.load_workbook(os.path.join(project_dir, "data", file_name))

    def sheet_name_list(self):
        sheetnames_list = self.wb.sheetnames
        new_sheet_name = []
        for name in sheetnames_list:
            if "test" in name:
                new_sheet_name.append(name)
        # print(new_sheet_name)
        return new_sheet_name

    def sheet_data(self):
        sheetnames_list = self.sheet_name_list()
        cases_data = {}
        for sheet_name in sheetnames_list:
            ws = self.wb[sheet_name]
            values = list(ws.values)
            if values:
                file = values[0]
                cases_step = []
                for value in values[1:]:
                    case_dict = dict(zip(file, value))
                    case_dict[xls_head['added']] = add_parameter_data(case_dict)
                    case_list = parametric_case(case_dict)
                    cases_step = cases_step + case_list
                cases_data[sheet_name] = cases_step
        return cases_data

    def check_session(self, sheet_name):
        ws = self.wb['设置']
        values = list(ws.values)
        case_dict = {}
        if values:
            file = values[0]
            for value in values[1:]:
                case_dict = dict(zip(file, value))
                # print(case_dict)
        if case_dict:
            value = ast.literal_eval(case_dict['是否为sessinon方式'])
            try:
                result = value[sheet_name]
                return result
            except:
                return 'no'


def add_parameter_data(step_data):
    add_parameter_dict = {}
    if step_data[xls_head['added']]:
        add_parameter_dict = ast.literal_eval(step_data[xls_head['added']])
    if step_data[xls_head['headers']]:
        headers = {"headers": ast.literal_eval(step_data[xls_head['headers']])}
        add_parameter_dict.update(headers)
    # 请求体默认data传输（后续看是否需要通过headers识别）
    if step_data[xls_head['data']]:
        data = {"data": step_data[xls_head['data']]}
        add_parameter_dict.update(data)
    return add_parameter_dict


def parametric_case(case_dict):
    parametric_file_name = case_dict[xls_head['file']]
    if parametric_file_name:
        case_parametric_list = []
        parametric_data = []
        # 打开读取参数化内文件
        keys_file = open(os.path.join(project_dir, "data", parametric_file_name))
        keys = keys_file.read().splitlines()
        # keys = re.findall('(.*,.*)', keys)
        # print(keys)
        parametric_head = tuple(keys[0].split(','))
        # print(parametric_head)
        # # 参数化文件数据生成[{},{}……]格式数据
        for i in keys[1:]:
            value = tuple(i.split(','))
            # print(value)
            parametric_data.append(dict(zip(parametric_head, value)))

        n_step = 0
        for parametric_dict in parametric_data:
            case_str = str(case_dict)
            for key in parametric_dict.keys():
                case_str = case_str.replace("${" + key + "}", parametric_dict[key])
            case_transition_dict = ast.literal_eval(case_str)
            case_transition_dict[xls_head['name']] = case_transition_dict[xls_head['name']] + '__parametric_' + str(n_step)
            n_step += 1
            case_parametric_list.append(case_transition_dict)

        # print(parametric_data)
        return case_parametric_list
    else:
        return [case_dict]


def pick_up_value(extract_rule, response: requests.Response):
    extract_rule_dict = ast.literal_eval(extract_rule)
    rule_type = extract_rule_dict['type']
    place_keys = list(extract_rule_dict.keys())
    value_result = {}

    def re_rules(data_rules, data_str):
        for value_name in data_rules.keys():
            value = re.findall(data_rules[value_name], data_str)
            if value:
                value_result.update({value_name: value[0]})
            else:
                print(f"{value_name}正则无法提取出数据，请检查！")
                # value_result.update({value_name: None})

    def jsonpath_rules(data_rules, data_dict):
        for value_name in data_rules.keys():
            value = jsonpath.jsonpath(data_dict, data_rules[value_name])
            if value:
                value_result.update({value_name: value[0]})
            else:
                print(f"{value_name}正则无法提取出数据，请检查！")
                # value_result.update({value_name: None})

    if rule_type == 're':
        if 'response_body' in place_keys:
            re_rules(extract_rule_dict['response_body'], response.text)
        elif 'response_headers' in place_keys:
            re_rules(extract_rule_dict['response_headers'], str(response.headers))
        elif 'response_url' in place_keys:
            re_rules(extract_rule_dict['response_url'], response.url)
        else:
            print("正则提取暂时没有该规则提取数据！")
    if rule_type == 'jsonpath':
        if 'response_body' in place_keys:
            jsonpath_rules(extract_rule_dict['response_body'], response.json())
        if 'response_headers' in place_keys:
            jsonpath_rules(extract_rule_dict['response_headers'], response.headers)
        else:
            print("jsonpath提取暂时没有该规则提取数据！")
    # print(value_result)
    return value_result


def update_data(step_data, extract_values):
    # print(extract_values)
    step_data_str = str(step_data)
    # print(step_data_str)
    for value_name in extract_values:
        step_data_str = step_data_str.replace("#{" + value_name + "}", extract_values[value_name])
    result_dict = ast.literal_eval(step_data_str)
    return result_dict


if __name__ == '__main__':

    data = XlsxData().sheet_name_list()
    print(data)
